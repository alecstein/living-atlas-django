from django.http import HttpResponse, JsonResponse

# All lemma keys are of the form {lemma}@{group}@.
# All form keys are of the form {lemma}@{group}@{form}
# and have value {group}@{form}.
# The "@" symbol will not appear in any other key,
# and the "@" symbol will only appear in the value if the
# checkbox corresponds to a form

def render_to_carto_response(request):

    json_data = {}

    for key in request.POST:
        try:
            lemma_name, latin, homonym_id, group, form = key.split("@")
        except ValueError:
            continue
        json_data.setdefault(group, []).append(form)

    return JsonResponse(json_data, status=200)

from openpyxl import Workbook

def render_to_excel_response(request):

    response = HttpResponse()
    response['Content-Type'] = 'application\
                               /vnd.openxmlformats-officedocument.\
                               spreadsheetml.sheet'
    response['Content-Disposition'] = f'attachment; \
                                    filename=living-atlas.xlsx'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "living-atlas-export"

    columns = {'form'       : 1,
               'lemma'      : 2,
               'latin'      : 3,
               'homonym_id' : 4,
               'group'      : 5,}

    for col_name, col_idx in columns.items():
        cell = worksheet.cell(row = 1, column = col_idx)
        cell.value = col_name

    row_idx = 2
    for key in request.POST:
        try:
            lemma_name, latin, homonym_id, group, form = key.split("@")
        except ValueError:
            continue

        row_values = {'form'      : form,
                      'lemma'     : lemma, 
                      'latin'     : latin,
                      'homonym_id': homonym_id,
                      'group'     : group,}

        for col_name, row_value in row_values.items():
            cell = worksheet.cell(row = row_idx, column = columns[col_name])
            cell.value = row_value
        row_idx += 1

    workbook.save(response)
    return response