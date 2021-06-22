from django.shortcuts import render
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound
from django.shortcuts import redirect
from .models import Form
from openpyxl import Workbook
import ast

# Create your views here.

def ajax_view(request):
    """
    Server API endpoint for fetching data
    N: limit of number of queries to fetch
    queryset: models in the database, lazily evaluated
    """
    N = 10_000

    context = {}
    queryset = Form.objects

    if request.method == 'GET':

        query = request.GET.get('query')
        query_type = request.GET.get('type')
        group = request.GET.get('group')
        lang = request.GET.get('lang')
        form_filter = request.GET.get('form_filter')

        if query_type == 'list':
            items = set(query.split(' '))

            if lang == 'lemma':
                forms = queryset.filter(lemma__in = items)
            elif lang == 'latin':
                forms = queryset.filter(latin__in = items)

            found = forms.values_list(lang, flat = True)
            not_found = [item for item in items if item not in found]
            context['not_found'] = not_found

        elif query_type == 'regex':
            if lang == 'lemma':
                forms = queryset.filter(lemma__regex = f"{query}")
            if lang == 'latin':
                forms = queryset.filter(latin__regex = f"{query}")

        if form_filter:
            forms = forms.filter(form__regex = f"{form_filter}")

        if not forms:
            return HttpResponseNotFound("No results found")

        results = {}
        for form, lemma, latin in forms.values_list()[:N]:
            if results.get(lemma):
                results[lemma]['forms'].append(form)
            else:
                results[lemma] = {'latin':latin, 'forms':[form]}

        context['results'] = results
        context['group'] = group
        response = render(request, "query.html", context)

        response['Limit'] = N
        if len(forms) > N:
            response['Exceeds-Limit'] = len(forms)

        return response

def search_view(request):
    """
    The app revolves around this view, which is also the homepage.
    """
    if request.method == 'GET':
        return render(request, 'search.html')

    if request.method == 'POST':

        # All lemma keys are of the form {lemma}@{group}@.
        # All form keys are of the form {lemma}@{group}@{form}
        # and have value {group}@{form}.
        # The "@" symbol will not appear in any other key,
        # and the "@" symbol will only appear in the value if the
        # checkbox corresponds to a form

        post_to = request.POST['post_to']

        if post_to == "export":

            response = HttpResponse()
            response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response['Content-Disposition'] = f'attachment; filename=living-atlas.xlsx'

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "living-atlas-export"

            columns = {'form':1, 'lemma':2, 'latin':3, 'group':4}

            for col_name, col_idx in columns.items():
                cell = worksheet.cell(row = 1, column = col_idx)
                cell.value = col_name

            row_idx = 2
            for key in request.POST:
                if len(key.split("@")) != 4:
                    continue

                lemma, latin, group, form = key.split("@")
                row_values = {'form':form, 'lemma':lemma, 'latin':latin, 'group':group}

                for col_name, row_value in row_values.items():
                    cell = worksheet.cell(row = row_idx, column = columns[col_name])
                    cell.value = row_value
                row_idx += 1

            workbook.save(response)
            return response

        json_data = {}

        for key in request.POST:
            if "@" not in key: 
                continue
            value = request.POST[key]
            if "@" not in value:
                continue
            group, form = value.split("@")
            if json_data.get(group):
                json_data[group].append(form)
            else:
                json_data[group] = [form]

        response = JsonResponse(json_data, status=200)

        return response

 
def about_view(request):
    return render(request, 'about.html', {})

def howto_view(request):
    return render(request, 'howto.html', {})

