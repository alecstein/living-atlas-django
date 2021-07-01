from django.http import HttpResponse, JsonResponse
from time import time
import json

def timer(func):
    def wrap_func(*args, **kwargs):
        start = time()
        result = func(*args, **kwargs)
        end = time()
        print(f'Function {func.__name__!r} executed in {(end-start):.4f}s')
        return result
    return wrap_func

from openpyxl import Workbook

def render_to_excel_response(request):

    response = HttpResponse()
    response['Content-Type'] = 'application' \
                               '/vnd.openxmlformats-officedocument.' \
                               'spreadsheetml.sheet'
    response['Content-Disposition'] = 'attachment;' \
                                    'filename=living-atlas.xlsx'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "living-atlas-export"

    columns = {'form'       : 1,
               'lemma'      : 2,
               'latin'      : 3,
               'homonym_id' : 4,
               'group'      : 5,}

    for col_name, col_idx in columns.items():
        cell = worksheet.cell(row = 1, column = col_idx)
        cell.value = col_name

    row_idx = 2

    json_data = json.loads(request.body)
    forms = json_data['forms']
    for item in forms:
        item = json.loads(item)
        form = item['form']
        values = item['values']
        row_values = {'form'      : form,
                      'lemma'     : values['lemma'], 
                      'latin'     : values['latin'],
                      'homonym_id': values['homid'],
                      'group'     : values['group'],}

        for col_name, row_value in row_values.items():
            cell = worksheet.cell(row = row_idx, column = columns[col_name])
            cell.value = row_value
        row_idx += 1

    workbook.save(response)
    return response